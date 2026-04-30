import openpyxl
import sys

wb = openpyxl.load_workbook('/home/thien/chum/anhnguyet/2025河内周期明细表.xlsx', data_only=True)
print("=== SHEETS ===")
for name in wb.sheetnames:
    print(f"  '{name}'")

# Check the main sheet
target = None
for name in wb.sheetnames:
    if '明细' in name:
        target = name
        break
if not target:
    target = wb.sheetnames[0]

print(f"\n=== Using sheet: '{target}' ===\n")
ws = wb[target]
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")

# Print first 150 rows, first 10 columns
print("=== ROW SCAN (first 150 rows, cols A-J) ===")
for row_idx in range(1, min(ws.max_row + 1, 151)):
    vals = []
    for col_idx in range(1, 11):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is not None:
            vals.append(str(v)[:30])
        else:
            vals.append('')
    # Only print rows with at least one non-empty value
    if any(v for v in vals):
        print(f"R{row_idx:3d}: {' | '.join(vals)}")

# Also scan for month markers
print("\n=== MONTH MARKERS ===")
for row_idx in range(1, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=1).value
    if cell and ('月' in str(cell) or '每周' in str(cell) or '付' in str(cell) or '总' in str(cell) or '刀' in str(cell)):
        print(f"R{row_idx:3d}: {str(cell)[:80]}")
