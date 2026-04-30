import openpyxl

path = '/home/thien/chum/anhnguyet/anhnguyet241.github.io/Máy 001 tháng 3_2026.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
print("SHEETS:", wb.sheetnames)

ws = wb.active
print(f"Active Sheet: {ws.title}")

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=10, values_only=True)):
    vals = [str(v)[:20] if v is not None else '' for v in row]
    print(f"R{i+1}: {' | '.join(vals)}")
